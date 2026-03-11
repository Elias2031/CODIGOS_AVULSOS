import os
import zipfile
import xml.etree.ElementTree as ET
import pandas as pd
from datetime import datetime
import openpyxl

def jaccard_similarity(str1, str2):
    set1 = set(str1.split())
    set2 = set(str(str2).split())
    intersection = set1.intersection(set2)
    union = set1.union(set2)
    return len(intersection) / len(union)

diretorio_atual = os.getcwd()
arquivos_na_pasta = os.listdir(diretorio_atual)
arquivos_zip = [arquivo for arquivo in arquivos_na_pasta if arquivo.endswith('.zip')]
arquivo_excel = [arquivo for arquivo in arquivos_na_pasta if arquivo.endswith('.xlsx')]
icms_mensal = 0
contador = 2
qtd_notas = 0 

excel = pd.read_excel(arquivo_excel[0], sheet_name = 'APLICAÇÃO')

mes_planilha = input('Por favor digite o mês a ser alterado')
print(len(arquivo_excel))
if len(arquivo_excel) > 1 :
    planilha = openpyxl.load_workbook(arquivo_excel[0])
else: planilha = openpyxl.Workbook()

planilha_ativa = planilha.create_sheet(mes_planilha)

planilha_ativa['A1'] = "Chave da nota"
planilha_ativa['B1'] = "Valor Contabil"
planilha_ativa['C1'] = "Base ICMS"
planilha_ativa['D1'] = "Valor ICMS"



# VERIFICA SE HÁ PELO MENOS UM ARQUIVO ZIP NA PASTA
if not arquivos_zip:
    print("Nenhum arquivo ZIP encontrado na pasta.")
else:

    # LOOP PARA CADA ARQUIVO ZIP ENCONTRADO NA PASTA
    for nome_arquivo_zip in arquivos_zip:

        with zipfile.ZipFile(nome_arquivo_zip, 'r') as arquivo_zip:
            lista_de_arquivos = arquivo_zip.namelist()
              
            # LOOP PARA CADA ARQUIVO XML ENCONTRADO NO ARQUIVO ZIP 
            for arquivo_nome in lista_de_arquivos:
                if not arquivo_nome.startswith('CANCELADO'):
                    qtd_notas += 1 

                    # ABRE O ARQUIVO XML E EXTRAI VALORES DAS VARIAVEIS REFERENTE A NOTA FISCAL
                    with arquivo_zip.open(arquivo_nome) as arquivo_xml:
                        tree = ET.parse(arquivo_xml)
                        root = tree.getroot()
                        items = root.findall('.//{http://www.portalfiscal.inf.br/nfe}det')
                        modelo_nota = root.find('.//{http://www.portalfiscal.inf.br/nfe}mod').text if root.find('.//{http://www.portalfiscal.inf.br/nfe}mod') is not None else 0
                        serie_nota = root.find('.//{http://www.portalfiscal.inf.br/nfe}serie').text if root.find('.//{http://www.portalfiscal.inf.br/nfe}serie') is not None else 0
                        numero_nota = root.find('.//{http://www.portalfiscal.inf.br/nfe}nNF').text if root.find('.//{http://www.portalfiscal.inf.br/nfe}nNF') is not None else 0
                        chave_nota = root.find('.//{http://www.portalfiscal.inf.br/nfe}chNFe').text if root.find('.//{http://www.portalfiscal.inf.br/nfe}chNFe') is not None else 0
                        valor_total_nota = root.find('.//{http://www.portalfiscal.inf.br/nfe}vNF').text if root.find('.//{http://www.portalfiscal.inf.br/nfe}vNF') is not None else 0
                        valor_desc_nota = root.find('.//{http://www.portalfiscal.inf.br/nfe}vDesc').text if root.find('.//{http://www.portalfiscal.inf.br/nfe}vDesc') is not None else 0
                        valor_icms_nota = root.find('.//{http://www.portalfiscal.inf.br/nfe}vICMS').text if root.find('.//{http://www.portalfiscal.inf.br/nfe}vICMS') is not None else 0
                        data_emissao_nota = root.find('.//{http://www.portalfiscal.inf.br/nfe}dhEmi').text if root.find('.//{http://www.portalfiscal.inf.br/nfe}dhEmi') is not None else 0
                        data_objeto = data_obj = datetime.strptime(data_emissao_nota[:10], "%Y-%m-%d")
                        data_emissao_formatada_nota = data_objeto.strftime("%d%m%Y")
                        
                        
                        soma_icms_items = 0
                        soma_valor_items = 0
                        soma_base_icms_items = 0
                        # LOOP PARA CADA ITEM DENTRO DO ARQUIVO XML
                        for item in items:
                            melhor_similaridade = 0
                            melhor_correspondencia = None
                            
                            # EXTRAI OS VALORES DAS VARIAVEIS REFERENTES AOS ITEMS 
                            nome_item_nota = item.find('.//{http://www.portalfiscal.inf.br/nfe}xProd').text if item.find('.//{http://www.portalfiscal.inf.br/nfe}xProd') is not None else 0
                            ncm_item_nota = item.find('.//{http://www.portalfiscal.inf.br/nfe}NCM').text if item.find('.//{http://www.portalfiscal.inf.br/nfe}NCM') is not None else 0
                            cest_item_nota = item.find('.//{http://www.portalfiscal.inf.br/nfe}CEST').text if item.find('.//{http://www.portalfiscal.inf.br/nfe}CEST') is not None else 0
                            aliq_icms_item_nota = item.find('.//{http://www.portalfiscal.inf.br/nfe}pICMS').text if item.find('.//{http://www.portalfiscal.inf.br/nfe}pICMS') is not None else 0
                            valor_icms_item_nota = item.find('.//{http://www.portalfiscal.inf.br/nfe}vICMS').text if item.find('.//{http://www.portalfiscal.inf.br/nfe}pICMS') is not None else 0
                            codigo_item_nota = item.find('.//{http://www.portalfiscal.inf.br/nfe}cProd').text if item.find('.//{http://www.portalfiscal.inf.br/nfe}cProd') is not None else 0
                            valor_item_nota = item.find('.//{http://www.portalfiscal.inf.br/nfe}vProd').text if item.find('.//{http://www.portalfiscal.inf.br/nfe}vProd') is not None else 0
                            rbc_item_nota = item.find('.//{http://www.portalfiscal.inf.br/nfe}pRedBC').text if item.find('.//{http://www.portalfiscal.inf.br/nfe}pRedBC') is not None else 0
                            base_item_nota = item.find('.//{http://www.portalfiscal.inf.br/nfe}vBC').text if item.find('.//{http://www.portalfiscal.inf.br/nfe}vBC') is not None else 0
                            unidade_item_nota = item.find('.//{http://www.portalfiscal.inf.br/nfe}uCom').text if item.find('.//{http://www.portalfiscal.inf.br/nfe}uCom') is not None else 0
                            desconto_item_nota = item.find('.//{http://www.portalfiscal.inf.br/nfe}vDesc').text if item.find('.//{http://www.portalfiscal.inf.br/nfe}vDesc') is not None else 0
                            valor_item_nota =  float(valor_item_nota) -float(desconto_item_nota)
                            
                            
                            
                            
                            soma_valor_items += float(valor_item_nota)
                            soma_icms_items += float(valor_icms_item_nota)
                            soma_base_icms_items += float(base_item_nota)
                            
                            
                            
                            
                        planilha_ativa[f'A{contador}'] = chave_nota
                        planilha_ativa[f'B{contador}'] = soma_valor_items
                        planilha_ativa[f'C{contador}'] = soma_base_icms_items
                        planilha_ativa[f'D{contador}'] = soma_icms_items
                        contador += 1 

planilha_ativa[f'A{contador + 2 }'] = f'{qtd_notas} Notas'
planilha.save('comparar_dexion.xlsx')                            
                            
