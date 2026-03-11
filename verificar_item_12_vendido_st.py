import os
import zipfile
import xml.etree.ElementTree as ET 
import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

diretorio_atual = os.getcwd()
arquivos_entrada = os.listdir(os.path.join(diretorio_atual, 'NOTAS_ENTRADA'))
arquivos_zip_entradas = [arquivo for arquivo in arquivos_entrada if arquivo.endswith('.zip')]
arquivos_saida = os.listdir(os.path.join(diretorio_atual, 'NOTAS_SAIDA'))
arquivos_zip_saidas = [arquivo for arquivo in arquivos_saida if arquivo.endswith('.zip')]

notas_processadas = set()
nome_itens_comprados_a_12 = set()

def jaccard_similarity(str1, str2):
    # Convert to lowercase and remove punctuation
    str1 = re.sub(r'[^\w\s]', '', str1.lower())
    str2 = re.sub(r'[^\w\s]', '', str2.lower())
    
    set1 = set(str1.split())
    set2 = set(str2.split())
    
    intersection = set1.intersection(set2)
    union = set1.union(set2)
    
    return len(intersection) / len(union)

def formatar_planilha(arquivo):
    # Carregar a planilha
    wb = Workbook()
    ws = wb.active
    ws.title = "Itens Analisados"
    
    # Ler os dados do DataFrame
    df = pd.read_excel(arquivo)
    
    # Definir borda fina para as células
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Adicionar o cabeçalho
    for col_num, column_title in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_num, value=column_title)
        cell.font = Font(bold=True, color="FFFFFF")  # Texto em negrito e branco
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # Fundo azul
        cell.alignment = Alignment(horizontal='center')  # Centraliza o texto
        cell.border = thin_border  # Aplica borda às células do cabeçalho

    # Adicionar os dados
    for row_num, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=2):
        for col_num, value in enumerate(row, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(horizontal='center', vertical='center')  # Centraliza os dados horizontal e verticalmente
            cell.border = thin_border  # Aplica borda às células de dados
            
            # Formatar como moeda (se o valor for numérico)
            if isinstance(value, (int, float)):
                cell.number_format = 'R$ #,##0.00'  # Formato para dinheiro brasileiro

    # Formatar largura das colunas
    for col in range(1, len(df.columns) + 1):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = 20  # Define largura da coluna

    # Salvar a planilha formatada
    wb.save(arquivo)

def extrair_valores_notas(root, caminho, valor_padrao):
    """Extrai um valor de um elemento XML, retornando um padrão se não encontrado."""
    elemento = root.find(f'.//{{http://www.portalfiscal.inf.br/nfe}}{caminho}')
    return elemento.text if elemento is not None else valor_padrao

def extrair_dados_itens_entradas(items, chave_nota, uf_emitente, natureza_operacao):
    dados_itens = []
    for item in items:
        icms_item = item.find('.//{http://www.portalfiscal.inf.br/nfe}ICMS')
        aliq_icms_item = extrair_valores_notas(icms_item, 'pICMS', '0')
        base_icms = extrair_valores_notas(icms_item, 'vBC', '0')
        
        # Verifica se a alíquota é igual a 12%
        if aliq_icms_item == '12.00':
            nome_item_nota = extrair_valores_notas(item, 'xProd', 'ITEM IMPORTADO')
            ncm_item_nota = extrair_valores_notas(item, 'NCM', None)
            cfop_item = extrair_valores_notas(item, 'CFOP', None)
            valor_produto = extrair_valores_notas(item, 'vProd', None)
            
            nome_itens_comprados_a_12.add(nome_item_nota)
            
            # Adiciona o item somente se a alíquota for 12%
            dados_itens.append({
                'nome_item_nota': nome_item_nota,
                'cfop': cfop_item,
                'ncm': ncm_item_nota,
                'chave_nota': chave_nota,
                'aliq_icms': aliq_icms_item,
                'uf_destinatario': uf_emitente,  
                'valor_produto': valor_produto,
                'base_icms': base_icms,
                'natureza_operacao': natureza_operacao
            })
    
    return dados_itens

def processar_notas_entrada():
    itens_comprados_a_12 = []  # Lista para acumular os dados de todos os itens
    for arquivo_zip in arquivos_zip_entradas:
        caminho_zip = os.path.join('NOTAS_ENTRADA', arquivo_zip)
        with zipfile.ZipFile(caminho_zip, 'r') as zip_entrada:
            lista_de_arquivos = [f for f in zip_entrada.namelist() if not f.startswith('CANCELADO')]
            for arquivo_nome in lista_de_arquivos:
                with zip_entrada.open(arquivo_nome) as arquivo_xml:
                    tree = ET.parse(arquivo_xml)
                    root = tree.getroot()
                    chave_nota = extrair_valores_notas(root, 'chNFe', '0')
                    natureza_operacao = extrair_valores_notas(root, 'natOp', '0')
                    emitente = root.find('.//{http://www.portalfiscal.inf.br/nfe}emit') 
                    uf_emitente = extrair_valores_notas(emitente, 'UF', None)
                    
                    if chave_nota in notas_processadas:
                        continue
                    else:
                        notas_processadas.add(chave_nota)
                    
                    itens = root.findall('.//{http://www.portalfiscal.inf.br/nfe}det')
                    dados_itens_filtrados = extrair_dados_itens_entradas(itens, chave_nota, uf_emitente, natureza_operacao)
                    itens_comprados_a_12.extend(dados_itens_filtrados)
    
    df = pd.DataFrame(itens_comprados_a_12)
    df.to_excel("itens_aliquota_12.xlsx", index=False)
    # formatar_planilha("itens_aliquota_12.xlsx")
    processar_notas_saida(itens_comprados_a_12)

def extrair_dados_itens_saidas(items, chave_nota, uf_emitente, uf_destinatario):
    dados_itens = []
    for item in items:
        nome_item_nota = extrair_valores_notas(item, 'xProd', 'ITEM IMPORTADO')
        ncm_item_nota = extrair_valores_notas(item, 'NCM', None)
        cfop_item = extrair_valores_notas(item, 'CFOP', None)
        valor_produto = extrair_valores_notas(item, 'vProd', None)
        icms_item = item.find('.//{http://www.portalfiscal.inf.br/nfe}ICMS')
        aliq_icms_item = extrair_valores_notas(icms_item, 'pICMS', '0')
        base_icms = extrair_valores_notas(icms_item, 'vBC', '0')


        for item_comprado_a_12 in nome_itens_comprados_a_12:
            similaridade = jaccard_similarity(item_comprado_a_12, nome_item_nota)
            if similaridade > 0.7:
                dados_itens.append({
                'nome_item_nota': nome_item_nota,
                'cfop': cfop_item,
                'ncm': ncm_item_nota,
                'chave_nota': chave_nota,
                'aliq_icms': aliq_icms_item,
                'uf_emitente': uf_emitente,  
                'uf_destinatario': uf_destinatario,  
                'valor_produto': valor_produto,
                'base_icms': base_icms,
            })
    
    
    return dados_itens

def processar_notas_saida(itens_comprados_a_12):
    todos_itens_comprados_a_12_vendidos_st = []  # Initialize list for sold items

    for arquivo_zip in arquivos_zip_saidas:
        caminho_zip = os.path.join('NOTAS_SAIDA', arquivo_zip)
        with zipfile.ZipFile(caminho_zip, 'r') as zip_saida:
            lista_de_arquivos = [f for f in zip_saida.namelist() if not f.startswith('CANCELADO')]
            for arquivo_nome in lista_de_arquivos:
                with zip_saida.open(arquivo_nome) as arquivo_xml:
                    tree = ET.parse(arquivo_xml)
                    root = tree.getroot()
                    chave_nota = extrair_valores_notas(root, 'chNFe', '0')
                    
                    if chave_nota in notas_processadas:
                        continue
                    else:
                        notas_processadas.add(chave_nota)
                        
                    natureza_operacao = extrair_valores_notas(root, 'natOp', '0')
                    emitente = root.find('.//{http://www.portalfiscal.inf.br/nfe}emit') 
                    destinatario = root.find('.//{http://www.portalfiscal.inf.br/nfe}dest') 
                    uf_emitente = extrair_valores_notas(emitente, 'UF', None)
                    uf_destinatario = extrair_valores_notas(destinatario, 'UF', None)
                    
                    # Process only if the emitter and recipient states are different
                    if uf_emitente != uf_destinatario:                        
                        itens = root.findall('.//{http://www.portalfiscal.inf.br/nfe}det')
                        dados = extrair_dados_itens_saidas(itens, chave_nota, uf_emitente, uf_destinatario)
                        todos_itens_comprados_a_12_vendidos_st.extend(dados)
                    else:
                        continue

    # After processing all outgoing invoices, export the collected data to Excel
    if todos_itens_comprados_a_12_vendidos_st:
        df_saida = pd.DataFrame(todos_itens_comprados_a_12_vendidos_st)
        df_saida.to_excel("itens_aliquota_12_vendidos.xlsx", index=False)
        # formatar_planilha("itens_aliquota_12_vendidos.xlsx")
    else:
        print("Nenhum item vendido com alíquota de 12% foi encontrado.")

processar_notas_entrada()
