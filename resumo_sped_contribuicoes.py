#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Varre uma pasta com vários SPED Contribuições (.txt) de períodos diferentes,
extrai apuração PIS/COFINS (M200/M600 + detalhes M210/M220/M610/M620)
e gera UMA planilha Excel consolidada (.xlsx).

Saídas (no mesmo xlsx):
- Aba "Resumo": 1 linha por arquivo x tributo (PIS/COFINS) com totais
- Aba "Detalhes": linhas com M210/M220/M610/M620 (quando existirem)
- Aba "Arquivos": inventário dos arquivos processados e status

Uso:
python consolidar_sped_contrib.py "C:\\pasta\\speds_contrib" --out "C:\\saida\\consolidado.xlsx"

Requisitos:
pip install openpyxl
"""

from __future__ import annotations

import argparse
import json
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


# ---------------------------
# Helpers de SPED
# ---------------------------

def _safe_strip(s: str) -> str:
    return (s or "").strip()

def _split_sped(line: str) -> List[str]:
    return line.rstrip("\n").rstrip("\r").split("|")

def _record_code(line: str) -> Optional[str]:
    line = _safe_strip(line)
    if not line or "|" not in line:
        return None
    parts = line.split("|")
    if len(parts) < 3:
        return None
    return _safe_strip(parts[1]) or None

def _get(fields: List[str], idx: int) -> str:
    return fields[idx] if idx < len(fields) else ""


# ---------------------------
# Estruturas
# ---------------------------

@dataclass
class HeaderInfo:
    cnpj: str = ""
    nome_empresarial: str = ""
    uf: str = ""
    cod_mun: str = ""
    ind_sit_esp: str = ""
    dt_ini: str = ""
    dt_fin: str = ""
    ind_reg_cum: str = ""  # 0110

@dataclass
class ApuracaoResumo:
    tributo: str
    periodo_ini: str
    periodo_fim: str
    vl_tot_cont: str = ""
    vl_tot_cred_desc: str = ""
    vl_tot_cred_desc_ant: str = ""
    vl_tot_cont_rec: str = ""
    vl_tot_cont_rec_aj: str = ""
    vl_tot_cont_rec_dif: str = ""
    vl_tot_cont_rec_dif_ant: str = ""
    vl_tot_cont_rec_per: str = ""
    vl_cont_apur: str = ""
    vl_cont_dif: str = ""
    vl_cont_dif_ant: str = ""
    vl_cont_per: str = ""

@dataclass
class DetalheItem:
    tributo: str
    reg: str
    descricao: str
    codigo: str
    valor: str
    info_extra: Dict[str, Any]

@dataclass
class ParseResult:
    arquivo: str
    ok: bool
    erro: str
    header: HeaderInfo
    resumos: List[ApuracaoResumo]
    detalhes: List[DetalheItem]


# ---------------------------
# Parser (por arquivo)
# ---------------------------

class SpedContribParser:
    def __init__(self) -> None:
        self.header = HeaderInfo()
        self.resumos: List[ApuracaoResumo] = []
        self.detalhes: List[DetalheItem] = []
        self._dt_ini = ""
        self._dt_fin = ""

    def parse_path(self, path: Path) -> None:
        with path.open("r", encoding="utf-8", errors="replace") as f:
            for line in f:
                self._parse_line(line)

    def _parse_line(self, line: str) -> None:
        reg = _record_code(line)
        if not reg:
            return
        fields = _split_sped(line)

        if reg == "0000":
            # |0000|COD_VER|COD_FIN|DT_INI|DT_FIN|NOME|CNPJ|UF|COD_MUN|...
            self.header.dt_ini = _get(fields, 4)
            self.header.dt_fin = _get(fields, 5)
            self._dt_ini = self.header.dt_ini
            self._dt_fin = self.header.dt_fin
            self.header.nome_empresarial = _get(fields, 6)
            self.header.cnpj = _get(fields, 7)
            self.header.uf = _get(fields, 8)
            self.header.cod_mun = _get(fields, 9)
            self.header.ind_sit_esp = _get(fields, 10)

        elif reg == "0110":
            # |0110|COD_INC_TRIB|IND_APRO_CRED|COD_TIPO_CONT|IND_REG_CUM|
            self.header.ind_reg_cum = _get(fields, 5)

        elif reg == "M200":
            self.resumos.append(
                ApuracaoResumo(
                    tributo="PIS",
                    periodo_ini=self._dt_ini,
                    periodo_fim=self._dt_fin,
                    vl_tot_cont=_get(fields, 2),
                    vl_tot_cred_desc=_get(fields, 3),
                    vl_tot_cred_desc_ant=_get(fields, 4),
                    vl_tot_cont_rec=_get(fields, 5),
                    vl_tot_cont_rec_aj=_get(fields, 6),
                    vl_tot_cont_rec_dif=_get(fields, 7),
                    vl_tot_cont_rec_dif_ant=_get(fields, 8),
                    vl_tot_cont_rec_per=_get(fields, 9),
                    vl_cont_apur=_get(fields, 10),
                    vl_cont_dif=_get(fields, 11),
                    vl_cont_dif_ant=_get(fields, 12),
                    vl_cont_per=_get(fields, 13),
                )
            )

        elif reg == "M600":
            self.resumos.append(
                ApuracaoResumo(
                    tributo="COFINS",
                    periodo_ini=self._dt_ini,
                    periodo_fim=self._dt_fin,
                    vl_tot_cont=_get(fields, 2),
                    vl_tot_cred_desc=_get(fields, 3),
                    vl_tot_cred_desc_ant=_get(fields, 4),
                    vl_tot_cont_rec=_get(fields, 5),
                    vl_tot_cont_rec_aj=_get(fields, 6),
                    vl_tot_cont_rec_dif=_get(fields, 7),
                    vl_tot_cont_rec_dif_ant=_get(fields, 8),
                    vl_tot_cont_rec_per=_get(fields, 9),
                    vl_cont_apur=_get(fields, 10),
                    vl_cont_dif=_get(fields, 11),
                    vl_cont_dif_ant=_get(fields, 12),
                    vl_cont_per=_get(fields, 13),
                )
            )

        elif reg == "M210":
            self.detalhes.append(
                DetalheItem(
                    tributo="PIS",
                    reg="M210",
                    descricao="Detalhamento da apuração (por código de contribuição)",
                    codigo=_get(fields, 2),
                    valor=_get(fields, 8),
                    info_extra={
                        "vl_rec_brt": _get(fields, 3),
                        "vl_bc_cont": _get(fields, 4),
                        "aliq": _get(fields, 5),
                        "quant_bc": _get(fields, 6),
                        "aliq_quant": _get(fields, 7),
                    },
                )
            )

        elif reg == "M220":
            self.detalhes.append(
                DetalheItem(
                    tributo="PIS",
                    reg="M220",
                    descricao="Ajuste da contribuição apurada",
                    codigo=_get(fields, 4),
                    valor=_get(fields, 3),
                    info_extra={
                        "ind_aj": _get(fields, 2),
                        "num_doc": _get(fields, 5),
                        "descr_aj": _get(fields, 6),
                        "dt_ref": _get(fields, 7),
                    },
                )
            )

        elif reg == "M610":
            self.detalhes.append(
                DetalheItem(
                    tributo="COFINS",
                    reg="M610",
                    descricao="Detalhamento da apuração (por código de contribuição)",
                    codigo=_get(fields, 2),
                    valor=_get(fields, 8),
                    info_extra={
                        "vl_rec_brt": _get(fields, 3),
                        "vl_bc_cont": _get(fields, 4),
                        "aliq": _get(fields, 5),
                        "quant_bc": _get(fields, 6),
                        "aliq_quant": _get(fields, 7),
                    },
                )
            )

        elif reg == "M620":
            self.detalhes.append(
                DetalheItem(
                    tributo="COFINS",
                    reg="M620",
                    descricao="Ajuste da contribuição apurada",
                    codigo=_get(fields, 4),
                    valor=_get(fields, 3),
                    info_extra={
                        "ind_aj": _get(fields, 2),
                        "num_doc": _get(fields, 5),
                        "descr_aj": _get(fields, 6),
                        "dt_ref": _get(fields, 7),
                    },
                )
            )


def parse_one_file(path: Path) -> ParseResult:
    parser = SpedContribParser()
    try:
        parser.parse_path(path)
        return ParseResult(
            arquivo=str(path),
            ok=True,
            erro="",
            header=parser.header,
            resumos=parser.resumos,
            detalhes=parser.detalhes,
        )
    except Exception as e:
        return ParseResult(
            arquivo=str(path),
            ok=False,
            erro=str(e),
            header=HeaderInfo(),
            resumos=[],
            detalhes=[],
        )


# ---------------------------
# Excel
# ---------------------------

def _autosize(ws) -> None:
    for col in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col)
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=col).value
            if val is None:
                continue
            s = str(val)
            if len(s) > max_len:
                max_len = len(s)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

def write_excel(out_xlsx: Path, results: List[ParseResult]) -> None:
    wb = Workbook()

    # Aba Resumo
    ws_res = wb.active
    ws_res.title = "Resumo"
    resumo_cols = [
        "arquivo",
        "cnpj", "nome_empresarial", "uf",
        "dt_ini", "dt_fin",
        "ind_reg_cum",
        "tributo",
        "vl_tot_cont", "vl_tot_cred_desc", "vl_tot_cred_desc_ant",
        "vl_tot_cont_rec", "vl_tot_cont_rec_aj", "vl_tot_cont_rec_dif",
        "vl_tot_cont_rec_dif_ant", "vl_tot_cont_rec_per",
        "vl_cont_apur", "vl_cont_dif", "vl_cont_dif_ant", "vl_cont_per",
    ]
    ws_res.append(resumo_cols)

    # Aba Detalhes
    ws_det = wb.create_sheet("Detalhes")
    det_cols = [
        "arquivo",
        "cnpj",
        "dt_ini", "dt_fin",
        "tributo",
        "reg",
        "descricao",
        "codigo",
        "valor",
        "info_extra_json",
    ]
    ws_det.append(det_cols)

    # Aba Arquivos
    ws_arq = wb.create_sheet("Arquivos")
    ws_arq.append(["arquivo", "ok", "erro"])

    for r in results:
        ws_arq.append([r.arquivo, r.ok, r.erro])

        if not r.ok:
            continue

        # Resumo: 1 linha por resumo (M200/M600)
        for s in r.resumos:
            ws_res.append([
                r.arquivo,
                r.header.cnpj,
                r.header.nome_empresarial,
                r.header.uf,
                s.periodo_ini,
                s.periodo_fim,
                r.header.ind_reg_cum,
                s.tributo,
                s.vl_tot_cont,
                s.vl_tot_cred_desc,
                s.vl_tot_cred_desc_ant,
                s.vl_tot_cont_rec,
                s.vl_tot_cont_rec_aj,
                s.vl_tot_cont_rec_dif,
                s.vl_tot_cont_rec_dif_ant,
                s.vl_tot_cont_rec_per,
                s.vl_cont_apur,
                s.vl_cont_dif,
                s.vl_cont_dif_ant,
                s.vl_cont_per,
            ])

        # Detalhes
        for d in r.detalhes:
            ws_det.append([
                r.arquivo,
                r.header.cnpj,
                r.header.dt_ini,
                r.header.dt_fin,
                d.tributo,
                d.reg,
                d.descricao,
                d.codigo,
                d.valor,
                json.dumps(d.info_extra, ensure_ascii=False),
            ])

    _autosize(ws_res)
    _autosize(ws_det)
    _autosize(ws_arq)

    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_xlsx)


# ---------------------------
# Varredura da pasta (FOR)
# ---------------------------

def find_sped_files(folder: Path) -> List[Path]:
    # Ajuste se você quiser incluir .sped etc.
    patterns = ["*.txt", "*.TXT"]
    files: List[Path] = []
    for pat in patterns:
        files.extend(folder.rglob(pat))
    # remove duplicados e ordena
    uniq = sorted(set(files))
    return uniq


def main() -> int:
    ap = argparse.ArgumentParser(description="Consolida vários SPED Contribuições em um único XLSX.")
    ap.add_argument("pasta", help="Pasta contendo os arquivos do SPED Contribuições")
    ap.add_argument("--out", default="consolidado_sped_contrib.xlsx", help="Caminho do XLSX de saída")
    args = ap.parse_args()

    pasta = Path(args.pasta).expanduser().resolve()
    out_xlsx = Path(args.out).expanduser().resolve()

    if not pasta.exists() or not pasta.is_dir():
        raise NotADirectoryError(f"Pasta inválida: {pasta}")

    arquivos = find_sped_files(pasta)
    if not arquivos:
        raise FileNotFoundError(f"Nenhum .txt encontrado em: {pasta}")

    results: List[ParseResult] = []
    for p in arquivos:  # <<<<<< FOR AQUI
        res = parse_one_file(p)
        results.append(res)

    write_excel(out_xlsx, results)
    print("[OK] XLSX gerado em:", out_xlsx)
    print("[INFO] Arquivos lidos:", len(arquivos))
    print("[INFO] OK:", sum(1 for r in results if r.ok), "| ERRO:", sum(1 for r in results if not r.ok))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
