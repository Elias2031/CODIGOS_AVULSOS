"""
Extrai APURAÇÃO de PIS e COFINS do SPED Contribuições (EFD PIS/COFINS) e gera .xlsx.

Este script foi ajustado para o layout típico encontrado no seu arquivo, onde:
- Apuração “oficial” está em M200 (PIS) e M600 (COFINS)
- Detalhamento de crédito por M100/M500 pode NÃO existir
- Informações de consolidação por CST/ALIQ aparecem em M210 (PIS) e M610 (COFINS)

Saída (.xlsx):
- Identificação: CNPJ, UF, DT_INI, DT_FIN, arquivo
- Apuração PIS (M200) e Apuração COFINS (M600)
- Auditoria: créditos efetivamente utilizados por M210/M610 (quando houver)
  (se não houver M210/M610, fica 0)

Uso:
  python extrair_pis_cofins_sped_contrib.py "C:\pasta\speds_contrib" "saida_pis_cofins.xlsx"
"""

from __future__ import annotations

import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, List, Tuple

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


# ---------------------------
# Helpers
# ---------------------------

def _to_decimal(v: str) -> float:
    v = (v or "").strip()
    if not v:
        return 0.0
    v = v.replace(",", ".")
    try:
        return float(v)
    except ValueError:
        return 0.0

def _try_read_text(path: Path) -> Tuple[str, str]:
    raw = path.read_bytes()
    for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
        try:
            return raw.decode(enc), enc
        except UnicodeDecodeError:
            continue
    return raw.decode("latin-1", errors="replace"), "latin-1(errors=replace)"


def _iter_sped_lines(text: str) -> Iterable[List[str]]:
    for line in text.splitlines():
        line = line.strip()
        if not line or "|" not in line:
            continue
        parts = line.split("|")
        # formato típico: ['', '0000', 'campo1', ..., '']
        if len(parts) >= 3 and parts[1]:
            yield parts


# ---------------------------
# Estrutura de saída
# ---------------------------

@dataclass
class SpedPisCofinsResumo:
    arquivo: str
    encoding: str

    cnpj: str = ""
    uf: str = ""
    dt_ini: str = ""
    dt_fin: str = ""

    # PIS (M200)
    pis_tot_cont_nc_per: float = 0.0
    pis_tot_cred_desc: float = 0.0
    pis_tot_cont_nc_dev: float = 0.0
    pis_cont_nc_rec: float = 0.0
    pis_tot_cont_cum_per: float = 0.0
    pis_cont_cum_rec: float = 0.0
    pis_tot_cont_rec: float = 0.0  # PIS a recolher

    # COFINS (M600)
    cofins_tot_cont_nc_per: float = 0.0
    cofins_tot_cred_desc: float = 0.0
    cofins_tot_cont_nc_dev: float = 0.0
    cofins_cont_nc_rec: float = 0.0
    cofins_tot_cont_cum_per: float = 0.0
    cofins_cont_cum_rec: float = 0.0
    cofins_tot_cont_rec: float = 0.0  # COFINS a recolher

    # Auditoria (quando existir): somatório de créditos informados na consolidação
    # M210: VL_CRED (na maioria dos layouts, campo 4 = parts[4])
    # M610: VL_CRED (na maioria dos layouts, campo 4 = parts[4])
    pis_creditos_consolidados_m210: float = 0.0
    cofins_creditos_consolidados_m610: float = 0.0

    linhas_processadas: int = 0


# ---------------------------
# Extração
# ---------------------------

def extrair_resumo_sped_contrib(path: Path) -> SpedPisCofinsResumo:
    text, enc = _try_read_text(path)
    resumo = SpedPisCofinsResumo(arquivo=path.name, encoding=enc)

    for parts in _iter_sped_lines(text):
        resumo.linhas_processadas += 1
        reg = parts[1].strip()

        # 0000 (EFD Contribuições)
        # |0000|...|DT_INI|DT_FIN|...|CNPJ|UF|...|
        # Exemplo típico: DT_INI=parts[6], DT_FIN=parts[7], CNPJ=parts[9], UF=parts[10]
        if reg == "0000":
            resumo.dt_ini = (parts[6] if len(parts) > 6 else "").strip()
            resumo.dt_fin = (parts[7] if len(parts) > 7 else "").strip()
            resumo.cnpj = (parts[9] if len(parts) > 9 else "").strip()
            resumo.uf = (parts[10] if len(parts) > 10 else "").strip()

        # M200 (Apuração PIS)
        # |M200|VL_TOT_CONT_NC_PER|VL_TOT_CRED_DESC|VL_TOT_CRED_DESC_ANT|VL_TOT_CONT_NC_DEV|...|VL_CONT_NC_REC|...|VL_TOT_CONT_CUM_PER|...|VL_CONT_CUM_REC|VL_TOT_CONT_REC|
        elif reg == "M200":
            resumo.pis_tot_cont_nc_per = _to_decimal(parts[2] if len(parts) > 2 else "")
            resumo.pis_tot_cred_desc = _to_decimal(parts[3] if len(parts) > 3 else "")
            resumo.pis_tot_cont_nc_dev = _to_decimal(parts[5] if len(parts) > 5 else "")
            resumo.pis_cont_nc_rec = _to_decimal(parts[8] if len(parts) > 8 else "")
            resumo.pis_tot_cont_cum_per = _to_decimal(parts[9] if len(parts) > 9 else "")
            resumo.pis_cont_cum_rec = _to_decimal(parts[12] if len(parts) > 12 else "")
            resumo.pis_tot_cont_rec = _to_decimal(parts[13] if len(parts) > 13 else "")

        # M600 (Apuração COFINS) — mesma estrutura do M200
        elif reg == "M600":
            resumo.cofins_tot_cont_nc_per = _to_decimal(parts[2] if len(parts) > 2 else "")
            resumo.cofins_tot_cred_desc = _to_decimal(parts[3] if len(parts) > 3 else "")
            resumo.cofins_tot_cont_nc_dev = _to_decimal(parts[5] if len(parts) > 5 else "")
            resumo.cofins_cont_nc_rec = _to_decimal(parts[8] if len(parts) > 8 else "")
            resumo.cofins_tot_cont_cum_per = _to_decimal(parts[9] if len(parts) > 9 else "")
            resumo.cofins_cont_cum_rec = _to_decimal(parts[12] if len(parts) > 12 else "")
            resumo.cofins_tot_cont_rec = _to_decimal(parts[13] if len(parts) > 13 else "")

        # M210 (Consolidação PIS por CST/ALIQ) — para auditoria de crédito
        # Linha comum: |M210|CST_PIS|VL_REC_BRT|VL_BC_CONT|VL_CRED|...|
        # Aqui usamos VL_CRED como parts[4] (ajuste se necessário)
        elif reg == "M210":
            resumo.pis_creditos_consolidados_m210 += _to_decimal(parts[4] if len(parts) > 4 else "")

        # M610 (Consolidação COFINS por CST/ALIQ) — para auditoria de crédito
        # Linha comum: |M610|CST_COFINS|VL_REC_BRT|VL_BC_CONT|VL_CRED|...|
        elif reg == "M610":
            resumo.cofins_creditos_consolidados_m610 += _to_decimal(parts[4] if len(parts) > 4 else "")

    return resumo


# ---------------------------
# Arquivos
# ---------------------------

def listar_arquivos_sped_contrib(pasta: Path) -> List[Path]:
    exts = {".txt", ".efd", ".sped"}
    arquivos: List[Path] = []
    for p in pasta.rglob("*"):
        if not p.is_file():
            continue
        name_upper = p.name.upper()
        if p.suffix.lower() in exts or "SPED" in name_upper or "EFD" in name_upper or "PISCOFINS" in name_upper:
            arquivos.append(p)
    return sorted(arquivos)


# ---------------------------
# Excel
# ---------------------------

def salvar_xlsx(resumos: List[SpedPisCofinsResumo], output_xlsx: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Apuração PIS COFINS"

    # (atributo, nome_legivel)
    colunas = [
        ("arquivo", "Arquivo SPED"),
        ("encoding", "Codificação"),
        ("cnpj", "CNPJ"),
        ("uf", "UF"),
        ("dt_ini", "Período Inicial (DT_INI)"),
        ("dt_fin", "Período Final (DT_FIN)"),

        ("pis_tot_cont_rec", "PIS a Recolher (M200)"),
        ("pis_tot_cont_nc_per", "PIS Não Cumulativo – Total no Período (M200)"),
        ("pis_tot_cred_desc", "PIS – Créditos Descontados (M200)"),
        ("pis_tot_cont_nc_dev", "PIS Não Cumulativo – Devido (M200)"),
        ("pis_cont_nc_rec", "PIS Não Cumulativo – A Recolher (M200)"),
        ("pis_tot_cont_cum_per", "PIS Cumulativo – Total no Período (M200)"),
        ("pis_cont_cum_rec", "PIS Cumulativo – A Recolher (M200)"),

        ("cofins_tot_cont_rec", "COFINS a Recolher (M600)"),
        ("cofins_tot_cont_nc_per", "COFINS Não Cumulativo – Total no Período (M600)"),
        ("cofins_tot_cred_desc", "COFINS – Créditos Descontados (M600)"),
        ("cofins_tot_cont_nc_dev", "COFINS Não Cumulativo – Devido (M600)"),
        ("cofins_cont_nc_rec", "COFINS Não Cumulativo – A Recolher (M600)"),
        ("cofins_tot_cont_cum_per", "COFINS Cumulativo – Total no Período (M600)"),
        ("cofins_cont_cum_rec", "COFINS Cumulativo – A Recolher (M600)"),

        ("pis_creditos_consolidados_m210", "Auditoria – Créditos PIS Consolidados (M210)"),
        ("cofins_creditos_consolidados_m610", "Auditoria – Créditos COFINS Consolidados (M610)"),

        ("linhas_processadas", "Registros Processados"),
    ]

    ws.append([label for _, label in colunas])

    for r in resumos:
        ws.append([getattr(r, key) for key, _ in colunas])

    ws.freeze_panes = "A2"
    for col_idx, (_, label) in enumerate(colunas, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, min(60, len(label) + 2))

    wb.save(output_xlsx)


# ---------------------------
# Main
# ---------------------------

def main() -> int:
    if len(sys.argv) < 3:
        print("Uso: python extrair_pis_cofins_sped_contrib.py <pasta_sped_contrib> <saida.xlsx>")
        return 2

    pasta = Path(sys.argv[1]).expanduser().resolve()
    out = Path(sys.argv[2]).expanduser().resolve()

    if not pasta.exists() or not pasta.is_dir():
        print(f"Erro: pasta inválida: {pasta}")
        return 2

    arquivos = listar_arquivos_sped_contrib(pasta)
    if not arquivos:
        print("Nenhum SPED Contribuições encontrado. Ajuste o filtro em listar_arquivos_sped_contrib().")
        return 1

    resumos: List[SpedPisCofinsResumo] = []
    total = len(arquivos)

    for i, arq in enumerate(arquivos, start=1):
        print(f"[{i}/{total}] Processando: {arq.name}")
        try:
            resumos.append(extrair_resumo_sped_contrib(arq))
        except Exception as e:
            resumos.append(SpedPisCofinsResumo(arquivo=arq.name, encoding="erro", linhas_processadas=0))
            print(f"  ERRO: {e}")

    salvar_xlsx(resumos, out)
    print(f"OK: planilha gerada em: {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
