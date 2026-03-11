import xml.etree.ElementTree as ET
import os
import openpyxl  # type: ignore
from openpyxl import Workbook  # type: ignore
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment  # type: ignore

# Definição de estilos reutilizáveis
COR = 'F0F8FF'
BORDER_SIDE = Side(style='thin', color='FF000000')
THIN_BORDER = Border(left=BORDER_SIDE, right=BORDER_SIDE, top=BORDER_SIDE, bottom=BORDER_SIDE)
FILL_PATTERN = PatternFill(start_color=COR, end_color=COR, fill_type='solid')
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
HEADER_FONT = Font(size=12, bold=True)
CELL_FONT = Font(size=11)

def finalizar_planilha(dicionario_segregacao, dicionario_csts_originais, razao_social, periodo):
    global planilha_de_itens_ativa, planilha_de_itens
    # Inicia na linha seguinte à última preenchida
    row = planilha_de_itens_ativa.max_row + 1

    # Preencher dados do dicionário de segregação na coluna S (19) e T (20)
    for codigo, valor in dicionario_segregacao.items():
        planilha_de_itens_ativa.cell(row=row, column=19, value=codigo)
        if isinstance(valor, dict):
            valor_formatado = ", ".join(f"{k}:{v}" for k, v in valor.items())
            planilha_de_itens_ativa.cell(row=row, column=20, value=valor_formatado)
        else:
            planilha_de_itens_ativa.cell(row=row, column=20, value=valor)
        row += 1
    row += 1  # Espaço entre as seções

    # Cabeçalho para CSTS ORIGINAIS
    header_cell = planilha_de_itens_ativa.cell(row=row, column=19, value='CSTS ORIGINAIS')
    header_cell.fill = FILL_PATTERN
    header_cell.font = Font(bold=True, size=12)
    planilha_de_itens_ativa.cell(row=row, column=20).fill = FILL_PATTERN
    row += 1

    # Preencher dados do dicionário de CSTS originais
    for codigo, valor in dicionario_csts_originais.items():
        planilha_de_itens_ativa.cell(row=row, column=19, value=codigo)
        planilha_de_itens_ativa.cell(row=row, column=20, value=valor)
        row += 1

    # Centraliza o conteúdo de todas as células
    for row_cells in planilha_de_itens_ativa.iter_rows():
        for cell in row_cells:
            cell.alignment = CENTER_ALIGN

    os.makedirs('PLANILHAS DE SEGRECAÇÃO', exist_ok=True)
    planilha_de_itens.save(f'PLANILHAS DE SEGRECAÇÃO/SEGREGACAO_DE_RECEITAS_{razao_social}_{periodo}.xlsx')

def gerar_planilha_fisica():
    global planilha_de_itens, planilha_de_itens_ativa, contador
    contador = 2
    planilha_de_itens = Workbook()
    planilha_de_itens_ativa = planilha_de_itens.create_sheet('Planilha de itens')
    # Remove a planilha padrão, se existir
    if 'Sheet' in planilha_de_itens.sheetnames:
        planilha_de_itens.remove(planilha_de_itens['Sheet'])
    
    # Definição dos cabeçalhos (colunas 1 a 20)
    headers = [
        "NOME DO ITEM NOTA",    # 1
        "CHAVE NOTA",           # 2
        "CST ICMS NOTA",        # 3
        "CST PIS NOTA",         # 4
        "CST ICMS CORRETO",     # 5
        "CST PIS CORRETO",      # 6
        "NCM ITEM NOTA",        # 7
        "VALOR CONTABIL DO ITEM",  # 8
        "Valor do ICMS na nota",   # 9
        "DESCONTO DO ITEM",         # 10
        "VALOR DO FRETE DO ITEM",   # 11
        "OUTRO VALOR",              # 12
        "BASE DE CALCULO",          # 13
        "REDUÇÃO ICMS",             # 14
        "BASE DE CALCULO C/ REDUCAO", # 15
        "VALOR DO ICMS REAPURADO",    # 16
        "DATA DA NOTA",             # 17
        "",                         # 18 (vazio)
        "SEGREGAÇÃO DE RECEITAS",    # 19
        ""                          # 20 (vazio)
    ]
    widths = [51, 45, 15.60, 13.57, 19, 17, 17, 12.30, 12.30, 14.60, 16, 17, 15.71, 16, 10, 26.71, 10, 10, 10, 10]
    
    # Preencher os cabeçalhos e ajustar a largura das colunas
    for col, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = planilha_de_itens_ativa.cell(row=1, column=col, value=header)
        cell.fill = FILL_PATTERN
        cell.border = THIN_BORDER
        cell.font = HEADER_FONT
        col_letter = openpyxl.utils.get_column_letter(col)
        planilha_de_itens_ativa.column_dimensions[col_letter].width = width

def adicionar_item_na_planilha(item):
    global contador, planilha_de_itens_ativa
    # Valores de 1 a 17 correspondem aos cabeçalhos da planilha
    values = [
        item['nome_item_nota'],
        item['chave_nota'],
        item['cst_icms'],
        item['cst_pis_cofins'],
        item['cst_icms_correto'],
        item['cst_pis_cofins_correto'],
        item['ncm'],
        float(item['valor_contabil_item']),
        float(item['valor_icms_item_nota']),
        float(item['desconto_item_nota']),
        float(item['valor_frete_item']),
        float(item['outro_valor_item']),
        float(item['base_de_calculo_item_s_reducao']),
        float(item['reducao_base_calculo_excel']),
        float(item['base_de_calculo_item_c_reducao']),
        float(item['valor_icms_item_reapurado']),
        item['data']
    ]
    
    # Atribuição e formatação dos dados da linha
    for col, val in enumerate(values, start=1):
        cell = planilha_de_itens_ativa.cell(row=contador, column=col, value=val)
        cell.border = THIN_BORDER
        cell.font = CELL_FONT
        if isinstance(val, float):
            cell.number_format = '#,##0.0000'
    contador += 1

def gerar_planilhas_segregacao(dados_por_periodo_e_cnpj):
    for periodo in sorted(dados_por_periodo_e_cnpj.keys()):
        for cnpj, dados_empresa in dados_por_periodo_e_cnpj[periodo].items():
            gerar_planilha_fisica()
            for item in dados_empresa['itens']:
                adicionar_item_na_planilha(item)
            finalizar_planilha(
                dados_empresa['dicionario_csts_corretos'],
                dados_empresa['dicionario_csts_originais'],
                cnpj,
                periodo
            )
