"""
Extrai valores de ICMS declarados de arquivos SPED Fiscal (EFD ICMS/IPI) e
gera uma planilha .xlsx com:
- CNPJ, IE, UF, Período (DT_INI/DT_FIN)
- Totais de ICMS por registro C190 (VL_ICMS, VL_ICMS_ST, VL_RED_BC)
- Totais de ICMS por registro C590 (VL_ICMS, VL_ICMS_ST, VL_RED_BC)
- Totais por E110 (VL_TOT_DEBITOS, VL_TOT_CREDITOS, VL_SLD_CREDOR_ANT,
  VL_SLD_APURADO, VL_TOT_DED, VL_ICMS_RECOLHER, VL_SLD_CREDOR_TRANSPORTAR,
  DEB_ESP)

Uso:
  python extrair_icms_sped.py "C:\pasta\speds" "saida_icms.xlsx"

Obs:
- SPED costuma ser "UTF-8" ou "latin-1". O script tenta detectar.
- Para “ICMS declarados” normalmente o registro mais “oficial” é o E110 (apuração).
  C190/C590 são somatórios por CFOP/CST/ALIQ; podem ajudar a auditar.
"""

from __future__ import annotations

import csv
import os
import sys
from dataclasses import dataclass, asdict
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, Optional, Tuple, List

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


# ---------------------------
# Helpers de parsing / números
# ---------------------------

def _to_decimal_str(v: str) -> float:
    """
    SPED usa decimal com ponto. Pode vir vazio.
    Retorna float (suficiente p/ planilha). Se quiser, troque por Decimal.
    """
    v = (v or "").strip()
    if not v:
        return 0.0
    # alguns arquivos podem vir com vírgula
    v = v.replace(",", ".")
    try:
        return float(v)
    except ValueError:
        return 0.0


def _try_read_text(path: Path) -> Tuple[str, str]:
    """
    Tenta ler arquivo como texto em encodings comuns de SPED.
    Retorna (texto, encoding_usado).
    """
    raw = path.read_bytes()
    for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
        try:
            return raw.decode(enc), enc
        except UnicodeDecodeError:
            continue
    # fallback “permissivo”
    return raw.decode("latin-1", errors="replace"), "latin-1(errors=replace)"


def _iter_sped_lines(text: str) -> Iterable[List[str]]:
    """
    Cada linha SPED: |REG|...|  (com pipe inicial e final).
    Faz split e remove vazios das pontas.
    """
    for line in text.splitlines():
        line = line.strip()
        if not line or "|" not in line:
            continue
        parts = line.split("|")
        # formato típico: ['', '0000', 'campo1', ..., '']
        if len(parts) >= 3 and parts[1]:
            yield parts


# ---------------------------
# Estruturas de saída
# ---------------------------

@dataclass
class SpedIcmsResumo:
    arquivo: str
    encoding: str

    cnpj: str = ""
    ie: str = ""
    uf: str = ""
    dt_ini: str = ""
    dt_fin: str = ""

    # Somatórios C190 (itens de documentos - analítico por CFOP/CST/ALIQ)
    c190_vl_icms: float = 0.0
    c190_vl_icms_st: float = 0.0
    c190_vl_red_bc: float = 0.0

    # Somatórios C590 (analítico serviços comunicação/telecom/energia etc)
    c590_vl_icms: float = 0.0
    c590_vl_icms_st: float = 0.0
    c590_vl_red_bc: float = 0.0

    # Apuração E110 (mais próximo de “declarado”)
    e110_vl_tot_debitos: float = 0.0
    e110_vl_tot_creditos: float = 0.0
    e110_vl_sld_credor_ant: float = 0.0
    e110_vl_sld_apurado: float = 0.0
    e110_vl_tot_ded: float = 0.0
    e110_vl_icms_recolher: float = 0.0
    e110_vl_sld_credor_transportar: float = 0.0
    e110_deb_esp: float = 0.0

    # Informativo
    linhas_processadas: int = 0


# ---------------------------
# Extração por arquivo
# ---------------------------

def extrair_resumo_sped(path: Path) -> SpedIcmsResumo:
    text, enc = _try_read_text(path)
    resumo = SpedIcmsResumo(arquivo=path.name, encoding=enc)

    for parts in _iter_sped_lines(text):
        resumo.linhas_processadas += 1
        reg = parts[1].strip()

        # 0000: abre arquivo / dados do contribuinte e período
        # Layout: |0000|COD_VER|COD_FIN|DT_INI|DT_FIN|NOME|CNPJ|CPF|UF|IE|COD_MUN|IM|SUFRAMA|IND_PERFIL|IND_ATIV|
        if reg == "0000":
            # índices conforme posição na linha split (parts[0] é '')
            resumo.dt_ini = (parts[4] if len(parts) > 4 else "").strip()
            resumo.dt_fin = (parts[5] if len(parts) > 5 else "").strip()
            resumo.cnpj = (parts[7] if len(parts) > 7 else "").strip()
            resumo.uf = (parts[9] if len(parts) > 9 else "").strip()
            resumo.ie = (parts[10] if len(parts) > 10 else "").strip()

        # C190: analítico do documento fiscal (ICMS)
        # |C190|CST_ICMS|CFOP|ALIQ_ICMS|VL_OPR|VL_BC_ICMS|VL_ICMS|VL_BC_ICMS_ST|VL_ICMS_ST|VL_RED_BC|VL_IPI|COD_OBS|
        elif reg == "C190":
            # VL_ICMS = parts[7], VL_ICMS_ST = parts[9], VL_RED_BC = parts[10]
            resumo.c190_vl_icms += _to_decimal_str(parts[7] if len(parts) > 7 else "")
            resumo.c190_vl_icms_st += _to_decimal_str(parts[9] if len(parts) > 9 else "")
            resumo.c190_vl_red_bc += _to_decimal_str(parts[10] if len(parts) > 10 else "")

        # C590: analítico serviços comunicação/telecom/energia etc.
        # |C590|CST_ICMS|CFOP|ALIQ_ICMS|VL_OPR|VL_BC_ICMS|VL_ICMS|VL_BC_ICMS_ST|VL_ICMS_ST|VL_RED_BC|COD_OBS|
        elif reg == "C590":
            resumo.c590_vl_icms += _to_decimal_str(parts[7] if len(parts) > 7 else "")
            resumo.c590_vl_icms_st += _to_decimal_str(parts[9] if len(parts) > 9 else "")
            resumo.c590_vl_red_bc += _to_decimal_str(parts[10] if len(parts) > 10 else "")

        # E110: apuração do ICMS
        # |E110|VL_TOT_DEBITOS|VL_AJ_DEBITOS|VL_TOT_AJ_DEBITOS|VL_ESTORNOS_CRED|
        #      |VL_TOT_CREDITOS|VL_AJ_CREDITOS|VL_TOT_AJ_CREDITOS|VL_ESTORNOS_DEB|
        #      |VL_SLD_CREDOR_ANT|VL_SLD_APURADO|VL_TOT_DED|VL_ICMS_RECOLHER|
        #      |VL_SLD_CREDOR_TRANSPORTAR|DEB_ESP|
        elif reg == "E110":
            # campos relevantes por posição:
            resumo.e110_vl_tot_debitos = _to_decimal_str(parts[2] if len(parts) > 2 else "")
            resumo.e110_vl_tot_creditos = _to_decimal_str(parts[6] if len(parts) > 6 else "")
            resumo.e110_vl_sld_credor_ant = _to_decimal_str(parts[10] if len(parts) > 10 else "")
            resumo.e110_vl_sld_apurado = _to_decimal_str(parts[11] if len(parts) > 11 else "")
            resumo.e110_vl_tot_ded = _to_decimal_str(parts[12] if len(parts) > 12 else "")
            resumo.e110_vl_icms_recolher = _to_decimal_str(parts[13] if len(parts) > 13 else "")
            resumo.e110_vl_sld_credor_transportar = _to_decimal_str(parts[14] if len(parts) > 14 else "")
            resumo.e110_deb_esp = _to_decimal_str(parts[15] if len(parts) > 15 else "")

    return resumo


# ---------------------------
# Varredura da pasta + Excel
# ---------------------------

def listar_arquivos_sped(pasta: Path) -> List[Path]:
    exts = {".txt", ".efd", ".sped"}  # ajuste se necessário
    arquivos: List[Path] = []
    for p in pasta.rglob("*"):
        if p.is_file():
            if p.suffix.lower() in exts or "SPED" in p.name.upper() or "EFD" in p.name.upper():
                arquivos.append(p)
    return sorted(arquivos)


def salvar_xlsx(resumos: List[SpedIcmsResumo], output_xlsx: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "ICMS Declarado"

    # (chave_dataclass, titulo_coluna_excel)
    colunas = [
        ("arquivo", "Arquivo SPED"),
        ("encoding", "Codificação do Arquivo"),

        ("cnpj", "CNPJ do Contribuinte"),
        ("ie", "Inscrição Estadual (IE)"),
        ("uf", "UF"),
        ("dt_ini", "Período Inicial da Apuração"),
        ("dt_fin", "Período Final da Apuração"),

        ("e110_vl_tot_debitos", "ICMS – Total de Débitos (E110)"),
        ("e110_vl_tot_creditos", "ICMS – Total de Créditos (E110)"),
        ("e110_vl_sld_credor_ant", "Saldo Credor Anterior de ICMS (E110)"),
        ("e110_vl_sld_apurado", "Saldo de ICMS Apurado no Período (E110)"),
        ("e110_vl_tot_ded", "Deduções do ICMS no Período (E110)"),
        ("e110_vl_icms_recolher", "ICMS a Recolher (E110)"),
        ("e110_vl_sld_credor_transportar", "Saldo Credor de ICMS a Transportar (E110)"),
        ("e110_deb_esp", "Débitos Especiais de ICMS (E110)"),

        ("c190_vl_icms", "ICMS Próprio – Mercadorias (C190)"),
        ("c190_vl_icms_st", "ICMS-ST – Mercadorias (C190)"),
        ("c190_vl_red_bc", "Redução de Base de Cálculo – Mercadorias (C190)"),

        ("c590_vl_icms", "ICMS Próprio – Serviços/Energia (C590)"),
        ("c590_vl_icms_st", "ICMS-ST – Serviços/Energia (C590)"),
        ("c590_vl_red_bc", "Redução de Base de Cálculo – Serviços/Energia (C590)"),

        ("linhas_processadas", "Quantidade de Registros Processados"),
    ]

    # Cabeçalho (rótulos)
    ws.append([label for _, label in colunas])

    # Linhas (valores pelos atributos)
    for r in resumos:
        ws.append([getattr(r, key) for key, _ in colunas])

    # Congelar cabeçalho
    ws.freeze_panes = "A2"

    # Ajuste de largura por tamanho do rótulo
    for col_idx, (_, label) in enumerate(colunas, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, min(55, len(label) + 2))

    wb.save(output_xlsx)
    
def main() -> int:
    if len(sys.argv) < 3:
        print("Uso: python extrair_icms_sped.py <pasta_sped> <saida.xlsx>")
        return 2

    pasta = Path(sys.argv[1]).expanduser().resolve()
    out = Path(sys.argv[2]).expanduser().resolve()

    if not pasta.exists() or not pasta.is_dir():
        print(f"Erro: pasta inválida: {pasta}")
        return 2

    arquivos = listar_arquivos_sped(pasta)
    if not arquivos:
        print("Nenhum arquivo SPED encontrado. Ajuste extensões/filtro em listar_arquivos_sped().")
        return 1

    resumos: List[SpedIcmsResumo] = []
    total = len(arquivos)

    for i, arq in enumerate(arquivos, start=1):
        print(f"[{i}/{total}] Processando: {arq.name}")
        try:
            resumos.append(extrair_resumo_sped(arq))
        except Exception as e:
            # registra “linha vazia” em caso de erro no arquivo
            resumos.append(SpedIcmsResumo(arquivo=arq.name, encoding="erro", linhas_processadas=0))
            print(f"  ERRO: {e}")

    salvar_xlsx(resumos, out)
    print(f"OK: planilha gerada em: {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
