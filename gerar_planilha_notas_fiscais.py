import os
import zipfile
import xml.etree.ElementTree as ET
from datetime import datetime
from itertools import count
from threading import Lock
from concurrent.futures import ThreadPoolExecutor
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Font

# Configurações de estilos para a planilha
cor = 'F0F8FF'
border1 = Side(style='thin', color='FF000000')
thin = Border(left=border1, right=border1, top=border1, bottom=border1)
fill = PatternFill(start_color=cor, end_color=cor, fill_type='solid')

contador = count(start=2)
contador_lock = Lock()

def extrair_valores_notas(root, caminho, valor_padrao):
    """Extrai um valor de um elemento XML, retornando um padrão se não encontrado."""
    elemento = root.find(f'.//{{http://www.portalfiscal.inf.br/nfe}}{caminho}')
    return elemento.text if elemento is not None else valor_padrao

def extrair_valor(root, caminho, valor_padrao):
    """Extrai um valor de um elemento XML, retornando um padrão se não encontrado."""
    elemento = root.find(f'.//{{http://www.portalfiscal.inf.br/nfe}}{caminho}')
    return elemento.text if elemento is not None else valor_padrao

def obter_contador():
    """Obtém o próximo valor do contador de forma thread-safe."""
    with contador_lock:
        return next(contador)

def gerar_planilha():
    """Cria uma nova planilha de itens e configura os cabeçalhos."""
    global planilha_de_itens, planilha_de_itens_ativa
    planilha_de_itens = openpyxl.Workbook()
    planilha_de_itens_ativa = planilha_de_itens.create_sheet('Planilha de itens')
    planilha_de_itens.remove(planilha_de_itens['Sheet'])  # Remove a sheet padrão criada

    headers = ["Nome do item", "NCM", "Chave da Nota", "Número da Nota", "Data de Emissão", 
               "Valor do Item", "Tipo de Nota", "CNPJ Emitente", "Razão Social Emitente", "Valor Base de calculo Item", "CFOP"]
    widths = [51, 45, 35, 15, 20, 20, 20, 20, 20, 20, 20]

    for col, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = planilha_de_itens_ativa.cell(row=1, column=col, value=header)
        cell.fill = fill
        cell.border = thin
        cell.font = Font(size=12, bold=True)
        planilha_de_itens_ativa.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

def adicionar_item_na_planilha(item):
    """Adiciona um item à planilha."""
    linha = item['contador']
    planilha_de_itens_ativa[f'A{linha}'] = item['nome_item_nota']
    planilha_de_itens_ativa[f'B{linha}'] = item['ncm_item_nota']
    planilha_de_itens_ativa[f'C{linha}'] = item['chave_nota']
    planilha_de_itens_ativa[f'D{linha}'] = item['numero_nota']
    planilha_de_itens_ativa[f'E{linha}'] = item['data_emissao_formatada_nota_para_excel']
    planilha_de_itens_ativa[f'F{linha}'] = item['valor_item_nota']
    planilha_de_itens_ativa[f'G{linha}'] = item['tipo_de_nota']
    planilha_de_itens_ativa[f'H{linha}'] = item['cnpj_emitente']
    planilha_de_itens_ativa[f'I{linha}'] = item['razao_social_emitente']
    planilha_de_itens_ativa[f'J{linha}'] = item['valor_base_de_calculo_item']
    planilha_de_itens_ativa[f'K{linha}'] = item['cfop']

    for coluna in "ABCDEFGHIJK":
        celula = planilha_de_itens_ativa[f"{coluna}{linha}"]
        celula.border = thin
        celula.font = Font(size=11)

def finalizar_planilha():
    """Salva a planilha criada no diretório 'planilhas'."""
    if not os.path.exists('planilhas'):
        os.makedirs('planilhas')
    planilha_de_itens.save('planilhas/planilha_de_itens.xlsx')

def processar_cte(root, arquivo_nome):
    """Processa o XML de um CT-e."""
    try:
        chave_cte = root.findtext('.//{http://www.portalfiscal.inf.br/cte}chCTe', '0')
        numero_cte = root.findtext('.//{http://www.portalfiscal.inf.br/cte}nCT', '0')
        data_emissao_cte = root.findtext('.//{http://www.portalfiscal.inf.br/cte}dhEmi', '0')
        namespaces = {"cte": "http://www.portalfiscal.inf.br/cte"}

        if data_emissao_cte != '0':
            data_objeto = datetime.strptime(data_emissao_cte[:10], "%Y-%m-%d")
            data_emissao_formatada_cte = data_objeto.strftime("%d/%m/%Y")
        else:
            data_emissao_formatada_cte = '0'

        emitente = root.find('.//{http://www.portalfiscal.inf.br/cte}emit')
        
        if emitente is not None:
            cnpj_emitente = emitente.findtext('{http://www.portalfiscal.inf.br/cte}CNPJ', '0')
            razao_social_emitente = emitente.findtext('{http://www.portalfiscal.inf.br/cte}xNome', '0')
        else:
            cnpj_emitente = '0'
            razao_social_emitente = 'Desconhecido'
        
        # Valor total da prestação do serviço
        valor_cte = float(root.findtext('.//cte:vTPrest', namespaces=namespaces) or 0)
        
        item_para_adicionar = {
                'nome_item_nota': 'NOTA CTE',
                'ncm_item_nota': 'NOTA CTE',
                'chave_nota': chave_cte,
                'numero_nota': numero_cte,
                'data_emissao_formatada_nota_para_excel': data_emissao_formatada_cte,
                'valor_item_nota': valor_cte,
                'contador': obter_contador(),
                'cnpj_emitente': cnpj_emitente,
                'razao_social_emitente': razao_social_emitente,
                'tipo_de_nota' : 'CTE'
            }

        adicionar_item_na_planilha(item_para_adicionar)

        # Adicione lógica para extrair os itens ou informações adicionais do CT-e, se necessário
        print(f"CT-e processado: Chave: {chave_cte}, Número: {numero_cte}, Emissão: {data_emissao_formatada_cte}")

    except Exception as e:
        print(f"Erro ao processar CT-e {arquivo_nome}: {e}")

def processar_nfe(root, arquivo_nome):
    """Processa o XML de uma NF-e."""
    itens_para_adicionar = []
    try:
        items = root.findall('.//{http://www.portalfiscal.inf.br/nfe}det')
        numero_nota = root.findtext('.//{http://www.portalfiscal.inf.br/nfe}nNF', '0')
        chave_nota = root.findtext('.//{http://www.portalfiscal.inf.br/nfe}chNFe', '0')
        data_emissao_nota = root.findtext('.//{http://www.portalfiscal.inf.br/nfe}dhEmi', '0')

        if data_emissao_nota != '0':
            data_objeto = datetime.strptime(data_emissao_nota[:10], "%Y-%m-%d")
            data_emissao_formatada_nota_para_excel = data_objeto.strftime("%d/%m/%Y")
        else:
            data_emissao_formatada_nota_para_excel = '0'

        emitente = root.find('.//{http://www.portalfiscal.inf.br/nfe}emit')
        if emitente is not None:
            cnpj_emitente = emitente.findtext('{http://www.portalfiscal.inf.br/nfe}CNPJ', '0')
            razao_social_emitente = emitente.findtext('{http://www.portalfiscal.inf.br/nfe}xNome', '0')
        else:
            cnpj_emitente = '0'
            razao_social_emitente = 'Desconhecido'

        for item in items:
            nome_item_nota = item.find('.//{http://www.portalfiscal.inf.br/nfe}xProd').text or 0
            print(nome_item_nota)
            ncm_item_nota = item.find('.//{http://www.portalfiscal.inf.br/nfe}NCM').text or 0
            valor_item_nota = item.find('.//{http://www.portalfiscal.inf.br/nfe}vProd').text or 0
            cfop = item.find('.//{http://www.portalfiscal.inf.br/nfe}CFOP').text or 0
            cst = item.find('.//{http://www.portalfiscal.inf.br/nfe}CFOP').text or 0
            valor_icms_item_nota = extrair_valores_notas(item, 'vICMS', 0)
            

            item_para_adicionar = {
                'nome_item_nota': nome_item_nota,
                'ncm_item_nota': ncm_item_nota,
                'chave_nota': chave_nota,
                'numero_nota': numero_nota,
                'data_emissao_formatada_nota_para_excel': data_emissao_formatada_nota_para_excel,
                'valor_item_nota': float(valor_item_nota),
                'contador': obter_contador(),
                'cnpj_emitente': cnpj_emitente,
                'razao_social_emitente': razao_social_emitente,
                'tipo_de_nota' : 'NFE',
                "valor_base_de_calculo_item": (
                    float(extrair_valor(item, 'vProd', None)) -
                    float(extrair_valor(item, 'vDesc', '0')) +
                    float(extrair_valor(item, 'vFrete', '0')) +
                    float(extrair_valor(item, 'vOutro', '0'))
                ),
                'cfop' : cfop,
                "ICMS" : valor_icms_item_nota
            }
            itens_para_adicionar.append(item_para_adicionar)

        for item in itens_para_adicionar:
            adicionar_item_na_planilha(item)

    except Exception as e:
        print(f"Erro ao processar NF-e {arquivo_nome}: {e}")

def identificar_tipo_documento(root):
    """Identifica se o documento XML é uma NF-e ou CT-e."""
    if root.find('.//{http://www.portalfiscal.inf.br/cte}infCte') is not None:
        return "CTe"
    elif root.find('.//{http://www.portalfiscal.inf.br/nfe}infNFe') is not None:
        return "NFe"
    return "Desconhecido"

def processar_xml(arquivo_zip, arquivo_nome):
    """Processa os arquivos XML dentro de um arquivo ZIP."""
    try:
        with arquivo_zip.open(arquivo_nome) as arquivo_xml:
            tree = ET.parse(arquivo_xml)
            root = tree.getroot()

            tipo_documento = identificar_tipo_documento(root)
            if tipo_documento == "NFe":
                processar_nfe(root, arquivo_nome)
            elif tipo_documento == "CTe":
                processar_cte(root, arquivo_nome)
            else:
                print(f"Tipo de documento desconhecido no arquivo {arquivo_nome}. Ignorado.")
    except Exception as e:
        print(f"Erro ao processar XML {arquivo_nome}: {e}")

def processar_zip(nome_arquivo_zip):
    """Processa os arquivos ZIP na pasta."""
    try:
        with zipfile.ZipFile(nome_arquivo_zip, 'r') as arquivo_zip:
            lista_de_arquivos = (nome for nome in arquivo_zip.namelist() if not nome.startswith('CANCELADO'))
            for arquivo_nome in lista_de_arquivos:
                processar_xml(arquivo_zip, arquivo_nome)
    except Exception as e:
        print(f"Erro ao processar ZIP {nome_arquivo_zip}: {e}")

if __name__ == "__main__":
    diretorio_atual = os.getcwd()
    arquivos_na_pasta = os.listdir(diretorio_atual)
    arquivos_zip = [arquivo for arquivo in arquivos_na_pasta if arquivo.endswith('.zip')]

    if not arquivos_zip:
        print("Nenhum arquivo ZIP encontrado na pasta.")
    else:
        gerar_planilha()
        with ThreadPoolExecutor() as executor:
            executor.map(processar_zip, arquivos_zip)
        finalizar_planilha()
